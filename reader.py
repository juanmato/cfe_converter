# reader.py — Lectura de archivos CFE en formato Excel (.xls / .xlsx)

import os
import logging
from datetime import datetime

from config import COLUMN_ALIASES

logger = logging.getLogger(__name__)


def _normalize(text):
    """Normaliza un texto para comparación flexible de nombres de columnas."""
    if text is None:
        return ""
    return str(text).strip().lower()


def _match_columns(header_row):
    """
    Dado un header del Excel, devuelve un dict {clave_interna: índice_columna}.
    Usa coincidencia flexible con COLUMN_ALIASES.
    """
    mapping = {}
    normalized_headers = [_normalize(h) for h in header_row]

    for key, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            for idx, header in enumerate(normalized_headers):
                if alias == header:
                    mapping[key] = idx
                    break
            if key in mapping:
                break

    return mapping


def _parse_fecha(valor):
    """Parsea una fecha desde string dd/mm/yyyy o desde datetime."""
    if isinstance(valor, datetime):
        return valor
    if valor is None:
        return None
    texto = str(valor).strip()
    if not texto or texto == "/  /":
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, fmt)
        except ValueError:
            continue
    return None


def _parse_monto(valor):
    """Convierte un valor a float. Retorna 0.0 si es None o no parseable."""
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip().replace(",", ".")
    if not texto or texto == "-":
        return 0.0
    try:
        return float(texto)
    except ValueError:
        return 0.0


def _parse_numero(valor):
    """Convierte el número de comprobante a string limpio."""
    if valor is None:
        return ""
    if isinstance(valor, (int, float)):
        return str(int(valor))
    return str(valor).strip()


def _parse_rut(valor):
    """Convierte el RUT a string limpio."""
    if valor is None:
        return ""
    return str(valor).strip()


def leer_excel(ruta_archivo):
    """
    Lee un archivo CFE en formato .xls o .xlsx.
    Retorna una lista de dicts con los campos normalizados.
    """
    ext = os.path.splitext(ruta_archivo)[1].lower()
    filas = []

    if ext == ".xlsx":
        filas = _leer_xlsx(ruta_archivo)
    elif ext == ".xls":
        filas = _leer_xls(ruta_archivo)
    else:
        raise ValueError(f"Formato no soportado: {ext}. Use .xls o .xlsx")

    if not filas:
        logger.warning("El archivo no contiene datos de CFE.")

    return filas


def _leer_xlsx(ruta):
    """Lee un archivo .xlsx con openpyxl. Busca en todas las hojas si la activa no tiene datos CFE."""
    import openpyxl

    wb = openpyxl.load_workbook(ruta, data_only=True)

    # Intentar primero la hoja activa, luego las demás
    hojas = [wb.active] + [wb[name] for name in wb.sheetnames if wb[name] != wb.active]

    for ws in hojas:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        registros = _procesar_filas(rows)
        if registros:
            logger.info(f"Datos CFE encontrados en hoja: '{ws.title}'")
            return registros

    # Ninguna hoja tuvo datos
    return _procesar_filas([])


def _leer_xls(ruta):
    """Lee un archivo .xls con xlrd. Busca en todas las hojas si la primera no tiene datos CFE."""
    import xlrd

    wb = xlrd.open_workbook(ruta)

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        rows = []
        for i in range(ws.nrows):
            rows.append(tuple(ws.cell_value(i, j) for j in range(ws.ncols)))
        if not rows:
            continue
        registros = _procesar_filas(rows)
        if registros:
            logger.info(f"Datos CFE encontrados en hoja: '{ws.name}'")
            return registros

    return _procesar_filas([])


def _encontrar_header(rows):
    """
    Busca la fila que contiene los headers de columnas CFE.
    Retorna (índice_fila, mapping) o (None, None) si no se encuentra.
    """
    for i, row in enumerate(rows):
        mapping = _match_columns(row)
        # Necesitamos al menos fecha, tipo, serie, numero, rut, moneda
        campos_requeridos = {"fecha_comprobante", "tipo_cfe", "serie", "numero", "rut_emisor", "moneda"}
        if campos_requeridos.issubset(mapping.keys()):
            return i, mapping
    return None, None


def _procesar_filas(rows):
    """Procesa las filas del Excel para extraer los registros CFE."""
    header_idx, mapping = _encontrar_header(rows)

    if header_idx is None:
        logger.error("No se encontró la fila de encabezados en el Excel.")
        return []

    logger.info(f"Encabezados encontrados en fila {header_idx + 1}: {mapping}")

    registros = []
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        if not row or all(v is None for v in row):
            continue

        # Extraer campo tipo_cfe para ver si es una fila de datos
        idx_tipo = mapping.get("tipo_cfe")
        tipo_val = row[idx_tipo] if idx_tipo is not None and idx_tipo < len(row) else None
        if tipo_val is None or str(tipo_val).strip() == "":
            continue

        fecha = _parse_fecha(row[mapping["fecha_comprobante"]] if mapping["fecha_comprobante"] < len(row) else None)
        if fecha is None:
            logger.warning(f"Fila {i + 1}: fecha inválida, se omite.")
            continue

        registro = {
            "fecha": fecha,
            "tipo_cfe": str(tipo_val).strip(),
            "serie": str(row[mapping["serie"]]).strip() if mapping.get("serie") is not None and mapping["serie"] < len(row) else "",
            "numero": _parse_numero(row[mapping["numero"]] if mapping.get("numero") is not None and mapping["numero"] < len(row) else None),
            "rut_emisor": _parse_rut(row[mapping["rut_emisor"]] if mapping.get("rut_emisor") is not None and mapping["rut_emisor"] < len(row) else None),
            "moneda": str(row[mapping["moneda"]]).strip() if mapping.get("moneda") is not None and mapping["moneda"] < len(row) else "",
            "monto_neto": _parse_monto(row[mapping["monto_neto"]] if mapping.get("monto_neto") is not None and mapping["monto_neto"] < len(row) else None),
            "iva_ventas": _parse_monto(row[mapping["iva_ventas"]] if mapping.get("iva_ventas") is not None and mapping["iva_ventas"] < len(row) else None),
            "monto_total": _parse_monto(row[mapping["monto_total"]] if mapping.get("monto_total") is not None and mapping["monto_total"] < len(row) else None),
            "monto_ret_per": _parse_monto(row[mapping.get("monto_ret_per", -1)] if mapping.get("monto_ret_per") is not None and mapping["monto_ret_per"] < len(row) else None),
            "monto_cred_fiscal": _parse_monto(row[mapping.get("monto_cred_fiscal", -1)] if mapping.get("monto_cred_fiscal") is not None and mapping["monto_cred_fiscal"] < len(row) else None),
        }

        registros.append(registro)

    return registros
